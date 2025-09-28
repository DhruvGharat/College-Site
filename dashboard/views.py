
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Avg, Max, Min, Count, F
from django.db import IntegrityError
import logging
import json

logger = logging.getLogger(__name__)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Faculty, Department, Subject, Student, Result, FacultySelection, PredefinedSubject
from django.db.models import Q
from django.db.utils import OperationalError

from .forms import FacultyLoginForm, FacultySelectionForm

from django.contrib.auth.decorators import login_required

def root_view(request):
    # Always start at a clean state: log out any existing session
    if request.user.is_authenticated:
        logout(request)
    return redirect('login')

@login_required
def subjectspage_view(request):
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')

    subjects = Subject.objects.filter(faculty=faculty).select_related('department').order_by('name')
    return render(request, 'dashboard/subjectspage.html', { 'subjects': subjects })


@login_required
def addsubjectpage_view(request):
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')

    if request.method == 'POST':
        dept_id = request.POST.get('department')
        year_input = request.POST.get('year')  # May now be a range like 2005-2006
        scheme = request.POST.get('scheme')
        name = request.POST.get('subjectName')
        code = request.POST.get('subjectCode')

        if not all([dept_id, year_input, scheme, name, code]):
            messages.error(request, 'All fields are required.')
            return redirect('addsubjectpage')

        try:
            department = Department.objects.get(id=int(dept_id))
        except (Department.DoesNotExist, ValueError, TypeError):
            messages.error(request, 'Invalid department selected.')
            return redirect('addsubjectpage')

        # Derive integer year from range if provided (fallback to 1)
        academic_year_val = ''
        try:
            if year_input and '-' in year_input:
                academic_year_val = year_input.strip()
                start_year = int(year_input.split('-')[0])
                # Map absolute start year crudely to relative year (1-4) if feasible
                # If not within a 4 year window, default to 1
                year_val = 1
            else:
                year_val = int(year_input)
                if year_val <= 0:
                    raise ValueError()
        except ValueError:
            messages.error(request, 'Invalid year / academic year value.')
            return redirect('addsubjectpage')

        # Normalize scheme values to match model choices
        scheme_map = {
            'R19-20': 'R19-20',
            'NEP': 'NEP',
            'AUTONOMOUS': 'AUTONOMOUS',
            'Autonomous': 'AUTONOMOUS',
            'R 19-20': 'R19-20',
        }
        scheme_val = scheme_map.get(scheme, scheme)

        try:
            # Always create a new subject instead of updating existing ones
            subject = Subject.objects.create(
                code=code,
                year=year_val,
                scheme=scheme_val,
                name=name,
                department=department,
                credits=3,
                faculty=faculty,
                academic_year=academic_year_val,
            )
            messages.success(request, 'Subject added successfully.')
            return redirect('subjectspage')
        except Exception as e:
            messages.error(request, f'Error saving subject: {str(e)}')
            return redirect('addsubjectpage')

    # GET
    allowed_department_names = [
        'Information Technology',
        'Electronics and telecommunication',
        'Computer Engineering',
    'Computer Science and Engineering ( IOT and Cybersecurity)'
    ]
    # Include legacy misspelling fallback so dropdown still shows department if rename migration not yet applied
    departments = Department.objects.filter(
        Q(name__in=allowed_department_names) |
        Q(name='Computer Science and Engineering ( IOT and Cybersrcurity)')
    ).order_by('name')
    # Build predefined subjects mapping { department_id: [names...] } with safe fallback if migration not applied
    predefined_map = {}
    try:
        for ps in PredefinedSubject.objects.filter(department__name='Information Technology', active=True).select_related('department'):
            predefined_map.setdefault(ps.department_id, []).append(ps.name)
    except OperationalError:
        # Table may not exist yet if migration not applied
        predefined_map = {}
    year_choices = Subject.YEAR_CHOICES
    scheme_choices = Subject.SCHEME_CHOICES
    return render(request, 'dashboard/addsubjectpage.html', {
        'departments': departments,
        'year_choices': year_choices,
        'scheme_choices': scheme_choices,
        'predefined_subjects_json': json.dumps(predefined_map),
    })


@login_required
def editsubjectpage_view(request, subject_id: int):
    # Ensure subject belongs to current faculty
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')

    subject = get_object_or_404(Subject, id=subject_id)
    if subject.faculty_id != faculty.id:
        messages.error(request, 'You do not have permission to edit this subject.')
        return redirect('subjectspage')

    if request.method == 'POST':
        dept_id = request.POST.get('department')
        year_input = request.POST.get('year')  # May be range now
        scheme = request.POST.get('scheme')
        name = request.POST.get('subjectName')
        code = request.POST.get('subjectCode')

        if not all([dept_id, year_input, scheme, name, code]):
            messages.error(request, 'All fields are required.')
            return redirect('editsubjectpage', subject_id=subject.id)

        try:
            department = Department.objects.get(id=int(dept_id))
        except (Department.DoesNotExist, ValueError, TypeError):
            messages.error(request, 'Invalid department selected.')
            return redirect('editsubjectpage', subject_id=subject.id)

        academic_year_val = ''
        try:
            if year_input and '-' in year_input:
                academic_year_val = year_input.strip()
                year_val = subject.year  # Keep prior integer year
            else:
                year_val = int(year_input)
                if year_val <= 0:
                    raise ValueError()
        except ValueError:
            messages.error(request, 'Invalid year / academic year value.')
            return redirect('editsubjectpage', subject_id=subject.id)

        scheme_map = {
            'R19-20': 'R19-20',
            'NEP': 'NEP',
            'AUTONOMOUS': 'AUTONOMOUS',
            'Autonomous': 'AUTONOMOUS',
            'R 19-20': 'R19-20',
        }
        scheme_val = scheme_map.get(scheme, scheme)

        try:
            # Apply updates
            subject.name = name
            subject.code = code
            subject.department = department
            subject.year = year_val
            subject.scheme = scheme_val
            subject.faculty = faculty
            if academic_year_val:
                subject.academic_year = academic_year_val
            subject.save()
            messages.success(request, 'Subject updated successfully.')
            return redirect('subjectspage')
        except IntegrityError:
            messages.error(request, 'A subject with this Code/Year/Scheme already exists.')
            return redirect('editsubjectpage', subject_id=subject.id)
        except Exception as e:
            messages.error(request, f'Error updating subject: {str(e)}')
            return redirect('editsubjectpage', subject_id=subject.id)

    # GET: render edit form with current values
    allowed_department_names = [
        'Information Technology',
        'Electronics and telecommunication',
        'Computer Engineering',
    'Computer Science and Engineering ( IOT and Cybersecurity)'
    ]
    departments = Department.objects.filter(
        Q(name__in=allowed_department_names) |
        Q(name='Computer Science and Engineering ( IOT and Cybersrcurity)')
    ).order_by('name')
    scheme_choices = Subject.SCHEME_CHOICES
    return render(request, 'dashboard/editsubjectpage.html', {
        'subject': subject,
        'departments': departments,
        'scheme_choices': scheme_choices,
    })


@login_required
@require_http_methods(["POST"])
def deletesubject_view(request, subject_id: int):
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')

    subject = get_object_or_404(Subject, id=subject_id)
    if subject.faculty_id != faculty.id:
        messages.error(request, 'You do not have permission to delete this subject.')
        return redirect('subjectspage')

    subject.delete()
    messages.success(request, 'Subject deleted successfully.')
    return redirect('subjectspage')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('subjectspage')
    
    if request.method == 'POST':
        print(f"POST request received: {request.POST}")  # Debug
        
        # Try direct authentication first (bypass form validation)
        username = request.POST.get('username')
        password = request.POST.get('password')
        print(f"Direct auth attempt: username='{username}', password='{password}'")  # Debug
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            print(f"User authenticated: {user}")  # Debug
            
            if user is not None:
                try:
                    faculty = Faculty.objects.get(user=user)
                    print(f"Faculty found: {faculty}")  # Debug
                    login(request, user)
                    print("Login successful, redirecting to subjectspage")  # Debug
                    return redirect('subjectspage')
                except Faculty.DoesNotExist:
                    print("Faculty.DoesNotExist error")  # Debug
                    messages.error(request, 'Access denied. Faculty account required.')
            else:
                print("Authentication failed")  # Debug
                messages.error(request, 'Invalid username or password.')
        else:
            print("Missing username or password")  # Debug
            messages.error(request, 'Please enter both username and password.')
        
        # Also try form validation for debugging
        form = FacultyLoginForm(request.POST)
        print(f"Form is valid: {form.is_valid()}")  # Debug
        if not form.is_valid():
            print(f"Form errors: {form.errors}")  # Debug
    else:
        form = FacultyLoginForm()
    
    return render(request, 'dashboard/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def selection_view(request):
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')
    
    if request.method == 'POST':
        form = FacultySelectionForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data['year']
            scheme = form.cleaned_data['scheme']
            department = form.cleaned_data['department']
            subject = form.cleaned_data.get('subject')
            
            # Save selection in session
            request.session['selected_year'] = year
            request.session['selected_scheme'] = scheme
            request.session['selected_department'] = department.id
            request.session['selected_subject'] = subject.id if subject else None
            
            # Save or update FacultySelection
            selection, created = FacultySelection.objects.get_or_create(
                faculty=faculty,
                year=year,
                scheme=scheme,
                department=department,
                defaults={'subject': subject}
            )
            if not created:
                selection.subject = subject
                selection.save()
            
            return redirect('dashboard')
    else:
        form = FacultySelectionForm()
    
    return render(request, 'dashboard/selection.html', {'form': form, 'faculty': faculty})


@login_required
def dashboard_view(request, subject_id=None):
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')
    
    # If subject_id is provided in URL, use it instead of session data
    if subject_id:
        try:
            selected_subject = Subject.objects.get(id=subject_id)
            # Update session with the selected subject
            request.session['selected_subject'] = selected_subject.id
            request.session['selected_year'] = selected_subject.year
            request.session['selected_scheme'] = selected_subject.scheme
            request.session['selected_department'] = selected_subject.department.id
        except Subject.DoesNotExist:
            messages.error(request, 'Subject not found.')
            return redirect('subjectspage')
    else:
        # Get session data
        selected_year = request.session.get('selected_year')
        selected_scheme = request.session.get('selected_scheme')
        selected_department_id = request.session.get('selected_department')
        selected_subject_id = request.session.get('selected_subject')
        
        # Get selected subject if available in session
        selected_subject = Subject.objects.get(id=selected_subject_id) if selected_subject_id else None
    
    # Determine display year preference: if a selected subject has an academic_year range use that.
    raw_year = request.session.get('selected_year')
    display_year = None
    if selected_subject and getattr(selected_subject, 'academic_year', ''):
        display_year = selected_subject.academic_year
    else:
        display_year = raw_year

    context = {
        'faculty': faculty,
        'selected_year': raw_year,  # keep original for any logic relying on numeric year
        'selected_display_year': display_year,
        'selected_scheme': request.session.get('selected_scheme'),
        'selected_department': Department.objects.get(id=request.session.get('selected_department')) if request.session.get('selected_department') else None,
        'selected_subject': selected_subject,
    }
    
    return render(request, 'dashboard/dashboard.html', context)


@login_required
def home_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
        
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')
    
    # Get session data
    selected_year = request.session.get('selected_year')
    selected_scheme = request.session.get('selected_scheme')
    selected_department_id = request.session.get('selected_department')
    selected_subject_id = request.session.get('selected_subject')
    
    # Motivational quotes
    quotes = [
        "Education is the most powerful weapon which you can use to change the world. - Nelson Mandela",
        "The future belongs to those who believe in the beauty of their dreams. - Eleanor Roosevelt",
        "Success is not final, failure is not fatal: it is the courage to continue that counts. - Winston Churchill",
        "The only way to do great work is to love what you do. - Steve Jobs",
        "Innovation distinguishes between a leader and a follower. - Steve Jobs"
    ]
    
    import random
    daily_quote = random.choice(quotes)
    
    context = {
        'faculty': faculty,
        'selected_year': selected_year,
        'selected_scheme': selected_scheme,
        'selected_department': Department.objects.get(id=selected_department_id) if selected_department_id else None,
        'selected_subject': Subject.objects.get(id=selected_subject_id) if selected_subject_id else None,
        'daily_quote': daily_quote,
    }
    
    return render(request, 'dashboard/home.html', context)


@login_required
def results_view(request):
    """Results Dashboard - Main results page with analytics and historical data"""
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')
    
    # Get session data
    selected_subject_id = request.session.get('selected_subject')
    selected_department_id = request.session.get('selected_department')
    
    current_subject = Subject.objects.get(id=selected_subject_id) if selected_subject_id else None
    current_department = Department.objects.get(id=selected_department_id) if selected_department_id else None
    
    # Initialize previous years data
    previous_years_data = []
    
    if current_subject:
        # Get current year and parse it
        current_year_str = str(current_subject.academic_year) if current_subject.academic_year else str(current_subject.year)
        try:
            # Parse year string like "2024-2025" or handle single year
            if '-' in current_year_str:
                current_start_year = int(current_year_str.split('-')[0])
            else:
                # Convert numeric year (1,2,3,4) to approximate academic year
                current_start_year = 2024  # Default to current academic year
            
            # Calculate previous two years (2023-2024, 2022-2023)
            for i in range(1, 3):
                prev_year_start = current_start_year - i
                prev_year_end = prev_year_start + 1
                prev_year_str = f"{prev_year_start}-{prev_year_end}"
                
                # Look for subjects with same code and department in previous years
                # Try multiple approaches to find historical data
                prev_subjects = []
                
                # 1. Look for subjects with academic_year field matching
                prev_subjects_with_academic_year = Subject.objects.filter(
                    code=current_subject.code,
                    department=current_subject.department,
                    academic_year=prev_year_str
                ).distinct()
                prev_subjects.extend(prev_subjects_with_academic_year)
                
                # 2. Look for subjects with same code/department that have results in that timeframe
                # (for data uploaded via Excel that might not have academic_year set properly)
                results_in_year = Result.objects.filter(
                    subject__code=current_subject.code,
                    subject__department=current_subject.department,
                    created_at__year__in=[prev_year_start, prev_year_end]  # Results created in those years
                ).values('subject').distinct()
                
                for result in results_in_year:
                    try:
                        subject = Subject.objects.get(id=result['subject'])
                        if subject not in prev_subjects:
                            prev_subjects.append(subject)
                    except Subject.DoesNotExist:
                        continue
                
                # Compile the data for this year
                year_data = {
                    'year': prev_year_str,
                    'subjects': [],
                    'has_data': False
                }
                
                # Add specific subject data if found
                for subject in prev_subjects:
                    # Get result statistics for this subject
                    subject_results = Result.objects.filter(subject=subject)
                    result_count = subject_results.count()
                    
                    if result_count > 0:
                        avg_marks = subject_results.aggregate(Avg('marks_obtained'))['marks_obtained__avg']
                        pass_count = subject_results.filter(marks_obtained__gte=40).count()
                        
                        year_data['subjects'].append({
                            'subject': subject,
                            'result_count': result_count,
                            'avg_marks': round(avg_marks, 2) if avg_marks else 0,
                            'pass_count': pass_count,
                            'pass_rate': round((pass_count / result_count) * 100, 2) if result_count > 0 else 0
                        })
                        year_data['has_data'] = True
                
                # Only show data for the selected subject, not department-wide subjects
                previous_years_data.append(year_data)
                
        except (ValueError, IndexError) as e:
            logger.error(f"Error parsing year data: {e}")
            
    elif current_department:
        # If only department is selected, show all subjects from that department in previous years
        current_year = 2024  # Default current year
        
        for i in range(1, 3):
            prev_year_start = current_year - i
            prev_year_end = prev_year_start + 1
            prev_year_str = f"{prev_year_start}-{prev_year_end}"
            
            # Get all subjects from this department with results in previous years
            dept_subjects = Subject.objects.filter(
                department=current_department
            ).filter(
                Q(academic_year=prev_year_str) |
                Q(result__created_at__year__in=[prev_year_start, prev_year_end])
            ).distinct()
            
            year_data = {
                'year': prev_year_str,
                'subjects': [],
                'has_data': False
            }
            
            for subject in dept_subjects:
                subject_results = Result.objects.filter(subject=subject)
                result_count = subject_results.count()
                
                if result_count > 0:
                    avg_marks = subject_results.aggregate(Avg('marks_obtained'))['marks_obtained__avg']
                    pass_count = subject_results.filter(marks_obtained__gte=40).count()
                    
                    year_data['subjects'].append({
                        'subject': subject,
                        'result_count': result_count,
                        'avg_marks': round(avg_marks, 2) if avg_marks else 0,
                        'pass_count': pass_count,
                        'pass_rate': round((pass_count / result_count) * 100, 2) if result_count > 0 else 0
                    })
                    year_data['has_data'] = True
            
            if year_data['has_data']:
                previous_years_data.append(year_data)
    
    context = {
        'faculty': faculty,
        'selected_subject': current_subject,
        'selected_department': current_department,
        'previous_years_data': previous_years_data,
    }
    
    return render(request, 'dashboard/results_dashboard.html', context)


@login_required
def subject_detail_view(request, subject_id, year=None):
    """Detailed view for a specific subject, optionally for a specific academic year"""
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')
    
    # Get the requested subject
    subject = get_object_or_404(Subject, id=subject_id)
    
    # If year is provided, try to find the subject for that specific year
    year_subject = None
    if year:
        try:
            # Try to find the same subject in the specified year
            year_subject = Subject.objects.get(
                code=subject.code,
                department=subject.department,
                academic_year=year
            )
        except Subject.DoesNotExist:
            # If not found, we'll still use the original subject but display data for the specified year
            year_subject = subject
    
    # Use the year-specific subject if found, otherwise use the original subject
    display_subject = year_subject if year_subject else subject
    
    # Get student results for this subject
    # If we're viewing a specific year, filter results by that year's subject
    # For the current year, just use the current subject
    if year and year_subject:
        # Get results for the specific year's subject
        results = Result.objects.filter(subject=year_subject).select_related('student').order_by('student__roll_number')
    elif year and not year_subject:
        # If we don't have a year-specific subject but have a year parameter,
        # try to find results for any subject with the same code in that academic year
        year_subjects = Subject.objects.filter(
            code=subject.code,
            academic_year=year
        )
        if year_subjects.exists():
            results = Result.objects.filter(subject__in=year_subjects).select_related('student').order_by('student__roll_number')
        else:
            # No subjects found for that year, show empty results
            results = Result.objects.none()
    else:
        # Default case - current year's subject
        results = Result.objects.filter(subject=display_subject).select_related('student').order_by('student__roll_number')
    
    # Calculate some basic analytics
    analytics = {}
    if results.exists():
        total_students = results.count()
        passing_students = results.filter(marks_obtained__gte=F('total_marks') * 0.4).count()  # Assuming 40% is passing
        
        analytics = {
            'total_students': total_students,
            'passing_students': passing_students,
            'passing_percentage': (passing_students / total_students * 100) if total_students > 0 else 0,
            'average_marks': results.aggregate(avg=Avg('marks_obtained'))['avg'],
            'highest_marks': results.aggregate(max=Max('marks_obtained'))['max'],
            'lowest_marks': results.aggregate(min=Min('marks_obtained'))['min'],
        }
    
    context = {
        'faculty': faculty,
        'subject': display_subject,
        'results': results,
        'year': year,  # Pass the year to the template
        'analytics': analytics,
        'is_previous_year': year is not None,
    }
    
    return render(request, 'dashboard/subject_detail.html', context)


# Placeholder views for other tabs
@login_required
def goal_set_view(request):
    return render(request, 'dashboard/coming_soon.html', {'title': 'Goal Set'})

@login_required
def tool_assignment_view(request):
    return render(request, 'dashboard/coming_soon.html', {'title': 'Tool Assignment'})

@login_required
def marks_entry_view(request):
    """Marks Entry page for entering student marks"""
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found.')
        return redirect('login')
    
    # Get session data
    selected_subject_id = request.session.get('selected_subject')
    selected_department_id = request.session.get('selected_department')
    selected_year = request.session.get('selected_year')
    selected_scheme = request.session.get('selected_scheme')
    
    if request.method == 'POST':
        # Handle marks submission
        try:
            subject = Subject.objects.get(id=selected_subject_id)
            department = Department.objects.get(id=selected_department_id)
            
            # Get students for this subject's criteria
            students = Student.objects.filter(
                department=department,
                year=subject.year,
                scheme=subject.scheme
            ).order_by('roll_number')
            
            saved_count = 0
            for student in students:
                marks_key = f'marks_{student.id}'
                total_key = f'total_{student.id}'
                exam_type_key = f'exam_type_{student.id}'
                
                if marks_key in request.POST and request.POST[marks_key]:
                    marks_obtained = int(request.POST[marks_key])
                    total_marks = int(request.POST.get(total_key, 100))
                    exam_type = request.POST.get(exam_type_key, 'Mid Term')
                    
                    # Create or update result
                    result, created = Result.objects.get_or_create(
                        student=student,
                        subject=subject,
                        exam_type=exam_type,
                        semester='1st',  # Default semester
                        defaults={
                            'marks_obtained': marks_obtained,
                            'total_marks': total_marks
                        }
                    )
                    
                    if not created:
                        result.marks_obtained = marks_obtained
                        result.total_marks = total_marks
                        result.save()
                    
                    saved_count += 1
            
            if saved_count > 0:
                messages.success(request, f'Successfully saved marks for {saved_count} students.')
            else:
                messages.warning(request, 'No marks were entered.')
                
        except (Subject.DoesNotExist, Department.DoesNotExist):
            messages.error(request, 'Selected subject or department not found.')
        except Exception as e:
            messages.error(request, f'Error saving marks: {str(e)}')
    
    # Get students for the selected criteria
    students = []
    if selected_subject_id and selected_department_id:
        try:
            subject = Subject.objects.get(id=selected_subject_id)
            department = Department.objects.get(id=selected_department_id)
            
            # Get students for this subject's criteria
            students = Student.objects.filter(
                department=department,
                year=subject.year,
                scheme=subject.scheme
            ).order_by('roll_number')
        except (Subject.DoesNotExist, Department.DoesNotExist):
            messages.warning(request, 'Selected subject or department not found.')
    
    context = {
        'faculty': faculty,
        'selected_subject': Subject.objects.get(id=selected_subject_id) if selected_subject_id else None,
        'selected_department': Department.objects.get(id=selected_department_id) if selected_department_id else None,
        'selected_year': selected_year,
        'selected_scheme': selected_scheme,
        'students': students,
    }
    
    return render(request, 'dashboard/marks_entry.html', context)

@login_required
def co_attainment_view(request):
    return render(request, 'dashboard/coming_soon.html', {'title': 'CO Attainment'})

@login_required
def co_po_mapping_view(request):
    return render(request, 'dashboard/coming_soon.html', {'title': 'CO-PO Mapping'})


@login_required
@require_http_methods(["GET"])
def download_marks_template(request):
    """Download Excel template for marks entry"""
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Marks Template"
        
        # Define headers
        headers = ['Roll No', 'Name', 'Marks Obtained', 'Total Marks', 'Exam Type']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add sample data
        sample_data = [
            ['21CS001', 'John Doe', 85, 100, 'Mid Term'],
            ['21CS002', 'Jane Smith', 92, 100, 'Mid Term'],
            ['21CS003', 'Bob Johnson', 78, 100, 'Mid Term'],
            ['21CS004', 'Alice Brown', 65, 100, 'Mid Term'],
            ['21CS005', 'Charlie Wilson', 88, 100, 'Mid Term']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="marks_template.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': f'Error creating template: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def upload_marks_excel(request):
    """Upload and parse Excel file with marks data"""
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty profile not found.'}, status=400)
    
    if 'excel_file' not in request.FILES:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    # Get session data
    selected_subject_id = request.session.get('selected_subject')
    if not selected_subject_id:
        return JsonResponse({'error': 'No subject selected. Please select a subject first.'}, status=400)
    
    try:
        subject = Subject.objects.get(id=selected_subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({'error': 'Selected subject not found.'}, status=400)
    
    try:
        # Read Excel file using openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        print(f"Excel file loaded successfully: {excel_file.name}")
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Validate headers
        required_headers = ['Roll No', 'Name', 'Marks Obtained', 'Total Marks', 'Exam Type']
        if not all(header in headers for header in required_headers):
            return JsonResponse({
                'error': f'Excel file must contain columns: {", ".join(required_headers)}'
            }, status=400)
        
        # Get column indices
        roll_no_idx = headers.index('Roll No')
        name_idx = headers.index('Name')
        marks_idx = headers.index('Marks Obtained')
        total_marks_idx = headers.index('Total Marks')
        exam_type_idx = headers.index('Exam Type')
        
        # Process data
        created_count = 0
        updated_count = 0
        errors = []
        uploaded_filename = excel_file.name
        
        print(f"Processing Excel file: {uploaded_filename}")
        print(f"Subject: {subject.name} (ID: {subject.id})")
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                roll_no = str(row[roll_no_idx]).strip() if row[roll_no_idx] else None
                name = str(row[name_idx]).strip() if row[name_idx] else None
                marks_obtained = int(row[marks_idx]) if row[marks_idx] else None
                total_marks = int(row[total_marks_idx]) if row[total_marks_idx] else 100
                exam_type = str(row[exam_type_idx]).strip() if row[exam_type_idx] else 'Mid Term'
                
                if not all([roll_no, name, marks_obtained is not None]):
                    errors.append(f'Row {row_num}: Missing required data')
                    continue
                
                # Get or create student
                student, student_created = Student.objects.get_or_create(
                    roll_number=roll_no,
                    defaults={
                        'name': name,
                        'department': subject.department,
                        'year': subject.year,
                        'scheme': subject.scheme
                    }
                )
                
                # Update student name if it has changed (even for existing students)
                if student.name != name:
                    student.name = name
                    student.save()
                    print(f"Updated student name: {roll_no} -> {name}")  # Debug log
                
                # Create or update result
                result, result_created = Result.objects.get_or_create(
                    student=student,
                    subject=subject,
                    exam_type=exam_type,
                    semester='1st',  # Default semester
                    defaults={
                        'marks_obtained': marks_obtained,
                        'total_marks': total_marks
                    }
                )
                
                if not result_created:
                    result.marks_obtained = marks_obtained
                    result.total_marks = total_marks
                    result.save()
                    updated_count += 1
                    print(f"Updated result for {student.roll_number}: {marks_obtained} marks")
                else:
                    created_count += 1
                    print(f"Created result for {student.roll_number}: {marks_obtained} marks")
                    
            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')
                continue
        
        # Store upload info in session
        request.session['last_uploaded_file'] = {
            'filename': uploaded_filename,
            'created': created_count,
            'updated': updated_count,
            'errors': len(errors)
        }
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully processed {created_count} new marks and updated {updated_count} existing marks from file: {uploaded_filename}',
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:10]  # Limit errors to first 10
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error processing Excel file: {str(e)}'}, status=400)


@login_required
@require_http_methods(["GET"])
def download_excel_template(request):
    """Download Excel template for results upload"""
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Results Template"
        
        # Define headers
        headers = ['Roll No', 'Name', 'Course Code', 'Marks']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add sample data
        sample_data = [
            ['21CS001', 'John Doe', 'CS201', 85],
            ['21CS002', 'Jane Smith', 'CS201', 92],
            ['21CS003', 'Bob Johnson', 'CS201', 78],
            ['21CS004', 'Alice Brown', 'CS201', 65],
            ['21CS005', 'Charlie Wilson', 'CS201', 88]
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="results_template.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': f'Error creating template: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def upload_excel_results(request):
    """Upload and parse Excel file with results"""
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        return JsonResponse({'error': 'Faculty profile not found.'}, status=400)
    
    if 'excel_file' not in request.FILES:
        return JsonResponse({'error': 'No file uploaded.'}, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        # Read Excel file using openpyxl
        from openpyxl import load_workbook
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info(f"Processing uploaded Excel file: {excel_file.name}, size: {excel_file.size} bytes")
        
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        logger.info(f"Found headers: {headers}")
        
        # Validate headers
        required_headers = ['Roll No', 'Name', 'Course Code', 'Marks']
        if not all(header in headers for header in required_headers):
            return JsonResponse({
                'error': f'Excel file must contain columns: {", ".join(required_headers)}'
            }, status=400)
        
        # Get column indices
        roll_no_idx = headers.index('Roll No')
        name_idx = headers.index('Name')
        course_code_idx = headers.index('Course Code')
        marks_idx = headers.index('Marks')
        
        # Pre-analyze data for summary statistics
        total_rows = 0
        valid_rows = 0
        total_marks = 0
        highest_mark = 0
        lowest_mark = 100
        course_data = {}
        
        # First pass - collect statistics
        for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
            try:
                total_rows += 1
                
                # Extract data
                roll_no = str(row[roll_no_idx].value).strip()
                name = str(row[name_idx].value).strip()
                course_code = str(row[course_code_idx].value).strip()
                
                # Validate marks
                try:
                    marks = float(row[marks_idx].value)
                    if marks < 0 or marks > 100:
                        errors.append(f'Row {row_num}: Marks must be between 0 and 100')
                        continue
                except (ValueError, TypeError):
                    errors.append(f'Row {row_num}: Invalid marks value')
                    continue
                
                valid_rows += 1
                total_marks += marks
                
                # Update highest and lowest marks
                highest_mark = max(highest_mark, marks)
                lowest_mark = min(lowest_mark, marks)
                
                # Collect course statistics
                if course_code not in course_data:
                    course_data[course_code] = {
                        'total': 0,
                        'sum': 0,
                        'pass': 0,
                        'fail': 0
                    }
                
                course_data[course_code]['total'] += 1
                course_data[course_code]['sum'] += marks
                
                if marks >= 40:
                    course_data[course_code]['pass'] += 1
                else:
                    course_data[course_code]['fail'] += 1
                    
            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')
                continue
        
        # Calculate summary statistics
        avg_marks = round(total_marks / valid_rows, 2) if valid_rows > 0 else 0
        pass_count = sum(data['pass'] for data in course_data.values())
        fail_count = sum(data['fail'] for data in course_data.values())
        
        # Process data for database
        created_count = 0
        updated_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                roll_no = str(row[roll_no_idx]).strip() if row[roll_no_idx] else None
                name = str(row[name_idx]).strip() if row[name_idx] else None
                course_code = str(row[course_code_idx]).strip() if row[course_code_idx] else None
                marks = int(row[marks_idx]) if row[marks_idx] else None
                
                if not all([roll_no, name, course_code, marks is not None]):
                    errors.append(f'Row {row_num}: Missing required data')
                    continue
                
                # Get or create student
                student, student_created = Student.objects.get_or_create(
                    roll_number=roll_no,
                    defaults={
                        'name': name,
                        'department': faculty.department,
                        'year': 2,  # Default year, can be made dynamic
                        'scheme': 'R19-20'  # Default scheme
                    }
                )
                
                # Update student name if it has changed (even for existing students)
                if student.name != name:
                    student.name = name
                    student.save()
                
                # Get subject
                try:
                    subject = Subject.objects.get(code=course_code)
                except Subject.DoesNotExist:
                    errors.append(f'Row {row_num}: Course code "{course_code}" not found')
                    continue
                
                # Create or update result
                result, result_created = Result.objects.get_or_create(
                    student=student,
                    subject=subject,
                    exam_type='Mid Term',
                    semester='1st',
                    defaults={'marks_obtained': marks}
                )
                
                if not result_created:
                    result.marks_obtained = marks
                    result.save()
                    updated_count += 1
                else:
                    created_count += 1
                    
            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')
                continue
        
        # Prepare course breakdown for response
        course_breakdown = []
        for code, data in course_data.items():
            try:
                subject = Subject.objects.get(code=code)
                course_name = subject.name
            except Subject.DoesNotExist:
                course_name = "Unknown Course"
                
            course_breakdown.append({
                'course_code': code,
                'course_name': course_name,
                'total_students': data['total'],
                'average_marks': round(data['sum'] / data['total'], 2) if data['total'] > 0 else 0,
                'pass_count': data['pass'],
                'fail_count': data['fail'],
                'pass_percentage': round((data['pass'] / data['total']) * 100, 2) if data['total'] > 0 else 0
            })
        
        # Return success response with analysis data
        return JsonResponse({
            'success': True,
            'message': f'Successfully processed {created_count} new results and updated {updated_count} existing results.',
            'created': created_count,
            'updated': updated_count,
            'errors': errors[:10],  # Limit errors to first 10
            'analysis': {
                'total_rows': total_rows,
                'valid_rows': valid_rows,
                'total_students': valid_rows,
                'pass_count': pass_count,
                'fail_count': fail_count,
                'average_marks': avg_marks,
                'highest_marks': highest_mark,
                'lowest_marks': lowest_mark,
                'pass_percentage': round((pass_count / valid_rows) * 100, 2) if valid_rows > 0 else 0,
                'per_course_breakdown': course_breakdown
            }
        })
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}", exc_info=True)
        return JsonResponse({'error': f'Error processing Excel file: {str(e)}'}, status=400)


@login_required
@require_http_methods(["POST"])
def remove_results_file(request):
    """Remove results file from database"""
    try:
        faculty = Faculty.objects.get(user=request.user)
        
        # Get selected subject
        selected_subject_id = request.session.get('selected_subject')
        if selected_subject_id:
            # Delete results for the selected subject
            deleted, _ = Result.objects.filter(subject_id=selected_subject_id).delete()
            return JsonResponse({'success': True, 'message': f'Removed {deleted} results from database'})
        else:
            return JsonResponse({'error': 'No subject selected'}, status=400)
            
    except Exception as e:
        logger.error(f"Error removing results: {str(e)}", exc_info=True)
        return JsonResponse({'error': f'Error removing results: {str(e)}'}, status=500)

@login_required
@require_http_methods(["GET"])
def download_analysis_excel(request):
    """Download analysis results as Excel file"""
    try:
        faculty = Faculty.objects.get(user=request.user)
        
        # Get selected subject or all subjects for the faculty's department
        selected_subject_id = request.session.get('selected_subject')
        subjects = (Subject.objects.filter(id=selected_subject_id) if selected_subject_id 
                   else Subject.objects.filter(department=faculty.department))
        
        if not subjects.exists():
            return HttpResponse('No subjects found', status=404)
            
        # Get results for the selected subjects
        results = Result.objects.filter(subject__in=subjects).select_related('student', 'subject')
        
        if not results.exists():
            return HttpResponse('No results found', status=404)
            
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Results Analysis"
        
        # Add headers
        headers = ['Subject', 'Total Students', 'Pass Count', 'Fail Count', 'Pass %', 'Average Marks', 'Highest Marks', 'Lowest Marks']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        # Add summary data per subject
        row = 2
        for subject in subjects:
            subject_results = results.filter(subject=subject)
            if subject_results.exists():
                total = subject_results.count()
                pass_count = subject_results.filter(marks_obtained__gte=40).count()
                fail_count = total - pass_count
                pass_percentage = (pass_count / total * 100) if total > 0 else 0
                avg_marks = subject_results.aggregate(Avg('marks_obtained'))['marks_obtained__avg'] or 0
                highest_marks = subject_results.aggregate(Max('marks_obtained'))['marks_obtained__max'] or 0
                lowest_marks = subject_results.aggregate(Min('marks_obtained'))['marks_obtained__min'] or 0
                
                data = [
                    subject.name,
                    total,
                    pass_count,
                    fail_count,
                    f"{pass_percentage:.2f}%",
                    f"{avg_marks:.2f}",
                    highest_marks,
                    lowest_marks
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                row += 1
                
        # Add overall summary
        row += 1
        ws.cell(row=row, column=1, value="Overall Summary").font = Font(bold=True)
        row += 1
        
        total = results.count()
        pass_count = results.filter(marks_obtained__gte=40).count()
        fail_count = total - pass_count
        pass_percentage = (pass_count / total * 100) if total > 0 else 0
        avg_marks = results.aggregate(Avg('marks_obtained'))['marks_obtained__avg'] or 0
        highest_marks = results.aggregate(Max('marks_obtained'))['marks_obtained__max'] or 0
        lowest_marks = results.aggregate(Min('marks_obtained'))['marks_obtained__min'] or 0
        
        summary_headers = ['Metric', 'Value']
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        row += 1
        
        summary_data = [
            ('Total Students', total),
            ('Pass Count', pass_count),
            ('Fail Count', fail_count),
            ('Pass Percentage', f"{pass_percentage:.2f}%"),
            ('Average Marks', f"{avg_marks:.2f}"),
            ('Highest Marks', highest_marks),
            ('Lowest Marks', lowest_marks)
        ]
        
        for item in summary_data:
            ws.cell(row=row, column=1, value=item[0])
            ws.cell(row=row, column=2, value=item[1])
            row += 1
            
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="results_analysis.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error creating analysis Excel: {str(e)}", exc_info=True)
        return HttpResponse(f'Error creating Excel file: {str(e)}', status=500)

@login_required
@require_http_methods(["GET"])
def download_student_results(request):
    """Download student results as Excel file"""
    try:
        faculty = Faculty.objects.get(user=request.user)
        
        # Get selected subject or all subjects for the faculty's department
        selected_subject_id = request.session.get('selected_subject')
        subjects = (Subject.objects.filter(id=selected_subject_id) if selected_subject_id 
                   else Subject.objects.filter(department=faculty.department))
        
        if not subjects.exists():
            return HttpResponse('No subjects found', status=404)
            
        # Get results for the selected subjects
        results = Result.objects.filter(subject__in=subjects).select_related('student', 'subject')
        
        if not results.exists():
            return HttpResponse('No results found', status=404)
            
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Results"
        
        # Add headers
        headers = ['Roll No', 'Student Name', 'Course Code', 'Course Name', 'Marks', 'Status', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        # Add student results
        for idx, result in enumerate(results, 2):
            row_data = [
                result.student.roll_number,
                result.student.name,
                result.subject.code,
                result.subject.name,
                result.marks_obtained,
                'Pass' if result.marks_obtained >= 40 else 'Fail',
                f"{result.percentage:.2f}%"
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=idx, column=col, value=value)
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_results.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error creating student results Excel: {str(e)}", exc_info=True)
        return HttpResponse(f'Error creating Excel file: {str(e)}', status=500)

@login_required
@require_http_methods(["GET"])
def results_analytics_api(request):
    """Get comprehensive results analytics"""
    try:
        logger.info(f"Starting results_analytics_api for user: {request.user.username}")
        
        # Get faculty profile
        faculty = Faculty.objects.filter(user=request.user).first()
        if not faculty:
            logger.error(f"Faculty profile not found for user {request.user.id}")
            return JsonResponse({'error': 'Faculty profile not found.'}, status=400)
            
        # Get selected subject or all subjects for the faculty's department
        selected_subject_id = request.session.get('selected_subject')
        subjects = (Subject.objects.filter(id=selected_subject_id) if selected_subject_id 
                   else Subject.objects.filter(department=faculty.department))
        
        # Even if no subjects exist, return a valid response structure
        # This ensures the frontend always has a consistent data format
        if not subjects.exists():
            logger.info(f"No subjects found for faculty {faculty.id}")
            return JsonResponse({
                'total_students': 0,
                'pass_count': 0,
                'fail_count': 0,
                'average_marks': 0,
                'highest_marks': 0,
                'lowest_marks': 0,
                'per_course_breakdown': [],
                'student_results': []
            }, status=200)
            
        # Get results for the selected subjects
        results = Result.objects.filter(subject__in=subjects).select_related('student', 'subject')
        
        if not results.exists():
            return JsonResponse({
                'total_students': 0,
                'pass_count': 0,
                'fail_count': 0,
                'average_marks': 0,
                'highest_marks': 0,
                'lowest_marks': 0,
                'per_course_breakdown': [],
                'student_results': []
            })
            
        # Calculate overall statistics
        total_students = results.count()
        pass_count = results.filter(marks_obtained__gte=40).count()
        
        # Get aggregates
        aggregates = results.aggregate(
            avg=Avg('marks_obtained'),
            max=Max('marks_obtained'),
            min=Min('marks_obtained')
        )
        
        # Prepare response
        response_data = {
            'total_students': total_students,
            'pass_count': pass_count,
            'fail_count': total_students - pass_count,
            'average_marks': round(aggregates['avg'], 2) if aggregates['avg'] is not None else 0,
            'highest_marks': aggregates['max'] if aggregates['max'] is not None else 0,
            'lowest_marks': aggregates['min'] if aggregates['min'] is not None else 0,
            'pass_percentage': round((pass_count / total_students) * 100, 2) if total_students > 0 else 0,
            'per_course_breakdown': [],
            'student_results': [{
                'roll_no': r.student.roll_number,
                'name': r.student.name,
                'course_code': r.subject.code,
                'course_name': r.subject.name,
                'marks': r.marks_obtained,
                'status': 'Pass' if r.marks_obtained >= 40 else 'Fail',
                'percentage': r.percentage
            } for r in results]
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"Error in results_analytics_api: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An error occurred while processing your request.'}, status=500)


@login_required
@require_http_methods(["GET"])
def subject_year_results_api(request, subject_id, year):
    """API endpoint to get detailed results for a specific subject and academic year"""
    try:
        faculty = Faculty.objects.get(user=request.user)
        
        # Get the subject
        subject = get_object_or_404(Subject, id=subject_id)
        
        # Parse the year parameter (should be in format "2023-2024")
        try:
            year_parts = year.split('-')
            if len(year_parts) == 2:
                year_start = int(year_parts[0])
                year_end = int(year_parts[1])
            else:
                raise ValueError("Invalid year format")
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid year format. Expected format: YYYY-YYYY'}, status=400)
        
        # Get results for this subject and year
        # Try multiple approaches to find results
        results_queryset = Result.objects.none()  # Start with empty queryset
        
        # 1. Look for subjects with matching academic_year
        subjects_with_academic_year = Subject.objects.filter(
            code=subject.code,
            department=subject.department,
            academic_year=year
        )
        if subjects_with_academic_year.exists():
            results_queryset = results_queryset.union(
                Result.objects.filter(subject__in=subjects_with_academic_year)
            )
        
        # 2. Look for results created in that time period
        results_by_date = Result.objects.filter(
            subject__code=subject.code,
            subject__department=subject.department,
            created_at__year__in=[year_start, year_end]
        )
        if results_by_date.exists():
            results_queryset = results_queryset.union(results_by_date)
        
        # Convert to list and get related data
        results = list(results_queryset.select_related('student', 'subject'))
        
        if not results:
            return JsonResponse({
                'success': False,
                'error': f'No results found for {subject.name} in academic year {year}'
            }, status=404)
        
        # Format the results data
        formatted_results = []
        for result in results:
            formatted_results.append({
                'student_roll': result.student.roll_number,
                'student_name': result.student.name,
                'marks_obtained': result.marks_obtained,
                'total_marks': result.total_marks,
                'status': result.status,
                'percentage': result.percentage,
                'exam_type': result.exam_type,
                'semester': result.semester,
                'created_at': result.created_at.strftime('%Y-%m-%d')
            })
        
        # Calculate summary statistics
        total_students = len(formatted_results)
        pass_count = sum(1 for r in formatted_results if r['status'] == 'Pass')
        avg_marks = sum(r['marks_obtained'] for r in formatted_results) / total_students if total_students > 0 else 0
        
        return JsonResponse({
            'success': True,
            'results': formatted_results,
            'summary': {
                'total_students': total_students,
                'pass_count': pass_count,
                'fail_count': total_students - pass_count,
                'pass_rate': round((pass_count / total_students) * 100, 2) if total_students > 0 else 0,
                'average_marks': round(avg_marks, 2),
                'subject_name': subject.name,
                'subject_code': subject.code,
                'academic_year': year
            }
        })
        
    except Faculty.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Faculty profile not found.'}, status=403)
    except Exception as e:
        logger.error(f"Error in subject_year_results_api: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'An error occurred while fetching results.'}, status=500)
